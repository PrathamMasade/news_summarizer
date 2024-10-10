import requests
from bs4 import BeautifulSoup
from selenium import webdriver
import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import google.generativeai as genai
import os
from dotenv import load_dotenv

load_dotenv('key.env')
api_key=os.getenv('api_key')

def gemini_summary(newstext,prompt):
    
    genai.configure(api_key=api_key)
    generation_config = {
        "temperature": 1,
        "top_p": 0.95,
        "top_k": 64,
        "max_output_tokens": 4000,
        "response_mime_type": "text/plain",
    }

    model = genai.GenerativeModel(model_name="gemini-1.5-pro", generation_config=generation_config)

    
    # Asking the model to extract the title, summary, and publication date
    chat_session = model.start_chat(history=[{"role": "user", "parts": [prompt]}])
    
    response = chat_session.send_message(newstext)
    
    return response.text

# Set up Selenium webdriver
driver = webdriver.Chrome()  # Replace with your preferred browser

# Send request to the URL and get the HTML response
url = "https://www.bbc.com/news/world-us-canada-64550143"

driver.get(url)
html = driver.page_source

# Parse the HTML content using BeautifulSoup
soup = BeautifulSoup(html, 'html.parser')

# Extract the article content
article_content = soup.find('p').text.strip()

# Load the T5 model and tokenizer
prompt="Give me the summary of this"
summary = gemini_summary(article_content,prompt)
title = soup.find('title').text.strip()
publication_date = soup.find('time').text.strip()
source = "BBC News"
url = url

# print(title)
# print(summary)

# Store the data in a dictionary
data = {
    'Title': title,
    'Summary': summary,
    'Publication Date': publication_date,
    'Source': source,
    'URL': url
}

#print(data)

# Write the data to a CSV file
wb = load_workbook("C:\\Users\\prath\\Downloads\\news_file.xlsx")
ws = wb.active
#print(ws.max_row)
next_row = ws.max_row + 1

ws[f'A{next_row}'] = data["Title"]
ws[f'B{next_row}'] = data["Summary"]
ws[f'C{next_row}'] = data["Publication Date"]
ws[f'D{next_row}'] = data["Source"]
ws[f'E{next_row}'] = data["URL"]

wb.save("C:\\Users\\prath\\Downloads\\news_file.xlsx")

# Close the Selenium webdriver
driver.quit()