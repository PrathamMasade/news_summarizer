from django.shortcuts import render
import requests
from django.contrib.auth.models import User
from django.http import JsonResponse
from django.shortcuts import render
from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework.authtoken.models import Token
import openpyxl

@api_view(['GET'])
def shownews(request):
    wb=openpyxl.load_workbook("C:\\Users\\prath\\Downloads\\news_file.xlsx")
    sheet = wb.active
    data=[]
    for r in sheet.iter_rows(values_only=True):
        data.append(list(r))
    return Response(data)

@api_view(['GET'])
def showbykey(request,key):
    wb=openpyxl.load_workbook("C:\\Users\\prath\\Downloads\\news_file.xlsx")
    sheet = wb.active
    data=[]
    for r in sheet.iter_rows(values_only=True):
        if r[5]==f"{key} \n":
            data.append(list(r))
    return Response(data)
