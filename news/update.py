import pandas as pd
import google.generativeai as genai
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
import os

load_dotenv('key.env')

api_key=os.getenv('api_key')

def gemini_summary(newstext,prompt):
    
    genai.configure(api_key=api_key)
    generation_config = {
        "temperature": 1,
        "top_p": 0.95,
        "top_k": 64,
        "max_output_tokens": 3000,
        "response_mime_type": "text/plain",
    }

    model = genai.GenerativeModel(model_name="gemini-1.5-pro", generation_config=generation_config)

    
    # Asking the model to extract the title, summary, and publication date
    chat_session = model.start_chat(history=[{"role": "user", "parts": [prompt]}])
    
    response = chat_session.send_message(newstext)
    
    return response.text

wb = load_workbook("C:\\Users\\prath\\Downloads\\news_file.xlsx")
ws = wb.active

prompt="categorize the news into one of the following categories and reply in just one word: politics,entertainment,environment,sports,others"
for i in range(1,ws.max_row+1):
    if ws[f'B{i}'].value!=None:
        ws[f'F{i}']= gemini_summary(ws[f'B{i}'].value,prompt)
    wb.save("C:\\Users\\prath\\Downloads\\news_file.xlsx")
